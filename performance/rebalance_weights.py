import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def rebalance_portfolio_weights():
    """
    Rebalances portfolio weights by adding new weights from the rebalance sheet
    to the existing weights file while preserving historical data.
    """
    
    # Get today's date
    today = datetime.now().strftime('%Y-%m-%d')
    print(f"Rebalancing portfolios for date: {today}")
    
    # Read the new weights from the rebalance sheet
    try:
        new_weights_df = pd.read_excel('new_weights.xlsx', sheet_name='rebalance')
        print(f"Successfully loaded rebalance sheet with {len(new_weights_df)} portfolios")
        
        # Debug: Show all columns in the rebalance sheet
        print(f"DEBUG - All columns in rebalance sheet: {list(new_weights_df.columns)}")
        print(f"DEBUG - Column types: {new_weights_df.dtypes.to_dict()}")
        
    except Exception as e:
        print(f"Error reading new_weights.xlsx: {e}")
        return
    
    # Load the existing weights Excel file
    try:
        weights_wb = openpyxl.load_workbook('weights.xlsx')
        print(f"Successfully loaded weights.xlsx with sheets: {weights_wb.sheetnames}")
    except Exception as e:
        print(f"Error loading weights.xlsx: {e}")
        return
    
    # Get portfolio names from the rebalance sheet
    portfolios = new_weights_df['Portfolio'].unique()
    
    # Get the correct date sequence from WC portfolio sheet for future dates
    future_dates = []
    try:
        if 'WC' in weights_wb.sheetnames:
            wc_df = pd.read_excel('weights.xlsx', sheet_name='WC')
            wc_df['Fecha'] = pd.to_datetime(wc_df['Fecha'])
            
            # Get dates after today
            today_dt = datetime.now()
            future_dates_from_wc = wc_df[wc_df['Fecha'] > today_dt]['Fecha'].sort_values()
            
            # Take the next 60 dates
            future_dates = future_dates_from_wc.head(60).dt.strftime('%Y-%m-%d').tolist()
            print(f"Found {len(future_dates)} future dates from WC sheet for extension")
            if future_dates:
                print(f"Date range: {future_dates[0]} to {future_dates[-1]}")
        else:
            print("WARNING: WC sheet not found, will only add today's date")
    except Exception as e:
        print(f"Error reading WC sheet dates: {e}")
        print("Will only add today's date")
    
    # If no future dates found, create a fallback (though this shouldn't happen with your setup)
    if not future_dates:
        future_dates = [today]
        print("Using only today's date as fallback")
    
    for portfolio in portfolios:
        print(f"\nProcessing portfolio: {portfolio}")
        
        # Get new weights for this portfolio
        portfolio_new_weights = new_weights_df[new_weights_df['Portfolio'] == portfolio].iloc[0]
        
        # Check if portfolio sheet exists in weights.xlsx
        if portfolio in weights_wb.sheetnames:
            # Load existing portfolio data
            existing_weights_df = pd.read_excel('weights.xlsx', sheet_name=portfolio)
            print(f"  Found existing sheet with {len(existing_weights_df)} historical records")
        else:
            # Create new portfolio sheet structure
            print(f"  Creating new sheet for portfolio: {portfolio}")
            existing_weights_df = pd.DataFrame({'Fecha': []})
            weights_wb.create_sheet(portfolio)
        
        # Prepare new rows with today's date and future dates
        new_rows = []
        
        # Get all asset columns from the rebalance data (excluding 'Portfolio' column and unnamed/nan columns)
        asset_columns = [col for col in new_weights_df.columns 
                        if col != 'Portfolio' 
                        and not str(col).startswith('Unnamed:')
                        and not str(col).startswith('nan.')
                        and pd.notna(col)
                        and str(col).strip() != '']
        
        print(f"DEBUG - All asset columns from rebalance: {len(asset_columns)} assets")
        
        # Create rows for today + next 60 business days
        dates_to_add = [today] + future_dates
        print(f"  Adding weights for {len(dates_to_add)} dates")
        
        assets_with_weights = 0
        for date in dates_to_add:
            new_row = {'Fecha': date}
            
            # Add weights for each asset (same weights for all future dates)
            for asset in asset_columns:
                weight_value = portfolio_new_weights[asset]
                # Add ALL assets to the new row, even if weight is 0 or NaN
                if pd.notna(weight_value) and weight_value != 0:
                    new_row[asset] = weight_value
                    if date == today:  # Only count once for logging
                        assets_with_weights += 1
                else:
                    new_row[asset] = 0  # Add with 0 weight instead of skipping
            
            new_rows.append(new_row)
        
        if dates_to_add:
            print(f"    Assets with non-zero weights: {assets_with_weights}")
            print(f"    Total assets per date: {len(asset_columns)}")
            print(f"    Date range: {dates_to_add[0]} to {dates_to_add[-1]}")
        
        # Add all new rows to existing data
        new_weights_df_multiple = pd.DataFrame(new_rows)
        updated_weights_df = pd.concat([existing_weights_df, new_weights_df_multiple], ignore_index=True)
        
        # Ensure ALL asset columns exist in the historical data (fill missing ones with NaN)
        for asset in asset_columns:
            if asset not in updated_weights_df.columns:
                updated_weights_df[asset] = np.nan
                print(f"    ADDED MISSING COLUMN: {asset}")
        
        # Add Total_Weight column for debugging (sum of all asset weights per row)
        # Get only asset columns (exclude 'Fecha' and any other non-asset columns)
        asset_cols_for_sum = [col for col in updated_weights_df.columns if col != 'Fecha' and col != 'Total_Weight']
        updated_weights_df['Total_Weight'] = updated_weights_df[asset_cols_for_sum].sum(axis=1, numeric_only=True)
        
        print(f"    ADDED Total_Weight column for debugging")
        
        # Remove duplicates based on 'Fecha' column (in case today's date already exists)
        updated_weights_df = updated_weights_df.drop_duplicates(subset=['Fecha'], keep='last')
        
        # Sort by date to maintain chronological order
        updated_weights_df['Fecha'] = pd.to_datetime(updated_weights_df['Fecha'])
        updated_weights_df = updated_weights_df.sort_values('Fecha')
        updated_weights_df['Fecha'] = updated_weights_df['Fecha'].dt.strftime('%Y-%m-%d')
        
        print(f"  Updated portfolio {portfolio} with new weights for {today}")
        print(f"  Total rows in updated sheet: {len(updated_weights_df)}")
        print(f"  Columns in updated sheet: {len(updated_weights_df.columns)} (including Total_Weight)")
        
        # Show total weights for the new dates added
        if len(dates_to_add) > 0:
            new_date_totals = updated_weights_df[updated_weights_df['Fecha'].isin(dates_to_add)]['Total_Weight']
            if len(new_date_totals) > 0:
                print(f"  Total weight for new dates: {new_date_totals.iloc[0]:.4f}")
        
        # Store the updated dataframe for later writing
        if not hasattr(rebalance_portfolio_weights, 'updated_portfolios'):
            rebalance_portfolio_weights.updated_portfolios = {}
        rebalance_portfolio_weights.updated_portfolios[portfolio] = updated_weights_df

    weights_wb.close()
    
    # Now write all updated portfolios back to Excel using pandas
    print(f"\nWriting updated data back to weights.xlsx...")
    
    with pd.ExcelWriter('weights.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        for portfolio, df in rebalance_portfolio_weights.updated_portfolios.items():
            df.to_excel(writer, sheet_name=portfolio, index=False)
            print(f"  Successfully wrote {len(df)} rows to sheet '{portfolio}'")
    
    # Save the updated weights file (overwriting the original)
    try:
        print(f"\nSuccessfully updated weights.xlsx with new portfolio weights for {today}")
    except Exception as e:
        print(f"Error saving file: {e}")

def display_summary():
    """
    Display a summary of the rebalancing operation
    """
    try:
        # Load both files for comparison
        new_weights_df = pd.read_excel('new_weights.xlsx', sheet_name='rebalance')
        
        print(f"\n{'='*50}")
        print("REBALANCING SUMMARY")
        print(f"{'='*50}")
        print(f"Date: {datetime.now().strftime('%Y-%m-%d')}")
        print(f"Portfolios processed: {len(new_weights_df)}")
        asset_columns_summary = [col for col in new_weights_df.columns 
                               if col != 'Portfolio' 
                               and not str(col).startswith('Unnamed:')
                               and not str(col).startswith('nan.')
                               and pd.notna(col)
                               and str(col).strip() != '']
        print(f"Valid assets in rebalance sheet: {len(asset_columns_summary)}")
        
        # Show portfolios and their total weight
        for _, row in new_weights_df.iterrows():
            portfolio = row['Portfolio']
            assets = {col: val for col, val in row.items() 
                     if col != 'Portfolio' 
                     and not str(col).startswith('Unnamed:')
                     and not str(col).startswith('nan.')
                     and pd.notna(col)
                     and str(col).strip() != ''
                     and pd.notna(val) and val != 0}
            total_weight = sum(assets.values())
            print(f"\n{portfolio}:")
            print(f"  Total weight: {total_weight:.4f}")
            print(f"  Number of assets: {len(assets)}")
            
    except Exception as e:
        print(f"Error generating summary: {e}")

if __name__ == "__main__":
    print("Portfolio Rebalancing Tool")
    print("=" * 30)
    
    # Run the rebalancing
    rebalance_portfolio_weights()
    
    # Display summary
    display_summary()
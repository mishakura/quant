#!/usr/bin/env python3
"""
Generate multiple Excel inputs from a generic Monte Carlo template.

Usage:
  python generate_excels.py

It loads `montecarlo/generic.xlsx` by default and writes files to
`montecarlo/generated/` named like `100.2000.M.xlsx`.

The script will try to auto-detect an "initial amount" cell (a header
containing the word "initial") and a "cashflow" sheet (name containing
"cash"). It writes the chosen initial amount and replaces cashflow
amounts in the detected column. You can customize paths and behavior
by editing the constants below or by adding CLI parsing.
"""
from pathlib import Path
import sys
try:
    import openpyxl
    from openpyxl.cell.cell import Cell
except Exception as e:
    print("openpyxl is required. Install with: pip install openpyxl")
    raise

TEMPLATE = Path(__file__).resolve().parent / "generic.xlsx"
OUT_DIR = Path(__file__).resolve().parent / "generated"

# Default cashflow list used for each initial amount
DEFAULT_CASHFLOWS = [100,200,300,400,500,600,700,800,900,1000,2000,3000,4000,5000]

# Initial amounts requested by the user (includes 100 and the extra list)
INITIAL_AMOUNTS = [100, 1000, 5000, 10000, 20000, 30000, 40000, 50000,60000,70000,80000,90000, 100000, 200000, 300000, 500000]


def find_header_cell(ws, keywords=('initial', 'initial amount')):
    """Search for a header cell containing one of keywords (case-insensitive).
    Returns (row, col) of the header cell or None."""
    kws = [k.lower() for k in keywords]
    for r in range(1, min(ws.max_row, 200) + 1):
        for c in range(1, min(ws.max_column, 50) + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and any(k in val.lower() for k in kws):
                return (r, c)
    return None


def set_initial_amount(wb, value):
    """Attempt to set the initial amount in the workbook.
    Strategy: look for a header containing 'initial' and set the cell to its right."""
    for name in wb.sheetnames:
        ws = wb[name]
        header = find_header_cell(ws)
        if header:
            r, c = header
            target = ws.cell(row=r, column=c+1)
            target.value = value
            return True
    return False


def find_cashflow_sheet(wb):
    for name in wb.sheetnames:
        if 'cash' in name.lower():
            return wb[name]
    # fallback: return sheet named 'cashflow' if exists
    if 'cashflow' in wb.sheetnames:
        return wb['cashflow']
    return None


def set_cashflow_amounts(wb, amount):
    ws = find_cashflow_sheet(wb)
    if ws is None:
        return False
    # Try to find an 'Amount' header
    header_col = None
    for r in range(1, min(ws.max_row, 50) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and 'amount' in val.lower():
                header_col = c
                start_row = r + 1
                break
        if header_col:
            break
    if header_col is None:
        # fallback: pick the first column that contains numeric data below row 1
        for c in range(1, min(ws.max_column, 20) + 1):
            for r in range(2, min(ws.max_row, 200) + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)):
                    header_col = c
                    start_row = 2
                    break
            if header_col:
                break
    if header_col is None:
        return False
    # write amount only into the first data row (the row immediately
    # after the header). Do not overwrite other rows.
    cell = ws.cell(row=start_row, column=header_col)
    cell.value = amount
    return True


def generate_files(template_path=TEMPLATE, out_dir=OUT_DIR,
                   initial_amounts=INITIAL_AMOUNTS,
                   cashflows=DEFAULT_CASHFLOWS):
    if not template_path.exists():
        print(f"Template not found: {template_path}")
        return
    out_dir.mkdir(parents=True, exist_ok=True)
    created = []
    for init in initial_amounts:
        for cf in cashflows:
            wb = openpyxl.load_workbook(template_path)
            ok_init = set_initial_amount(wb, init)
            ok_cf = set_cashflow_amounts(wb, cf)
            fname = f"{init}.{cf}.M.xlsx"
            path = out_dir / fname
            wb.save(path)
            created.append((path, ok_init, ok_cf))
    print(f"Created {len(created)} files in {out_dir}")
    # show a quick summary of detection success
    failures = [p for p in created if not (p[1] and p[2])]
    if failures:
        print("Warning: some files were created but detection/replace failed for template fields.")
        print("Open one file in Excel to inspect where the replacements should be made,")
        print("or customize the script by specifying the exact cell/column to edit.")


if __name__ == '__main__':
    generate_files()

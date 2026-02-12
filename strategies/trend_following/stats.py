import pandas as pd
import os
import numpy as np
import openpyxl
import yfinance as yf

from quant.utils.constants import TRADING_DAYS_PER_YEAR
from quant.analytics.performance import (
    annualized_return,
    annualized_volatility,
    max_drawdown,
    return_kurtosis,
    return_skewness,
    sharpe_ratio,
    win_rate as compute_win_rate,
)
from quant.data.providers.yfinance import YFinanceProvider

def calculate_daily_pnl(simulation_file='simulation_results.csv', main_file='main.csv', output_file='daily_pnl.csv'):
    # Load data
    sim_df = pd.read_csv(simulation_file, sep=',')  # Changed to comma separator
    main_df = pd.read_csv(main_file)

    # Debug: Print columns to verify
    print("Columns in sim_df:", sim_df.columns.tolist())
    print("Columns in main_df:", main_df.columns.tolist())

    # Convert dates (remove specific format to let pandas infer ISO format)
    sim_df['EntryDate'] = pd.to_datetime(sim_df['EntryDate'])
    sim_df['ExitDate'] = pd.to_datetime(sim_df['ExitDate'])
    main_df['Date'] = pd.to_datetime(main_df['Date'])  # Assuming main.csv dates are already parseable

    # Sort main_df for efficiency
    main_df = main_df.sort_values(['ticker', 'Date'])

    # Pre-group main_df by ticker for faster access
    ticker_data = {ticker: group for ticker, group in main_df.groupby('ticker')}

    daily_pnl_list = []
    total_trades = len(sim_df)

    for idx, trade in enumerate(sim_df.iterrows()):
        ticker = trade[1]['Ticker']  # trade is (index, row)
        entry_date = trade[1]['EntryDate']
        exit_date = trade[1]['ExitDate']
        usd_position = trade[1]['USD_Position']
        capital_before = trade[1]['Capital_before']  # Use capital for normalizing

        if ticker in ticker_data:
            # Filter the pre-grouped data for this ticker and date range, and valid reasons
            trade_days = ticker_data[ticker][
                (ticker_data[ticker]['Date'] >= entry_date) &
                (ticker_data[ticker]['Date'] <= exit_date) &
                (ticker_data[ticker]['Reason'].isin(['entry', 'hold', 'stop_loss', 'exit_min100']))
            ].sort_values('Date')

            if not trade_days.empty:
                # Calculate daily PnL contribution normalized by capital
                closes = trade_days['Close'].values
                dates = trade_days['Date'].values
                for i in range(1, len(closes)):
                    daily_return = (closes[i] - closes[i-1]) / closes[i-1]
                    daily_pnl_contrib = (usd_position * daily_return) / capital_before  # Normalize by capital
                    daily_pnl_list.append({'Date': dates[i], 'Daily_PnL_Contrib': daily_pnl_contrib})

        # Print progress
        progress = int((idx + 1) / total_trades * 100)
        print(f"Progress: {progress}%")

    # Aggregate by date: sum of normalized contributions
    if daily_pnl_list:
        daily_pnl_df = pd.DataFrame(daily_pnl_list)
        daily_agg = daily_pnl_df.groupby('Date').agg(
            Daily_PnL_Pct=('Daily_PnL_Contrib', 'sum')
        ).reset_index()
        daily_agg = daily_agg.sort_values('Date')
        daily_agg.to_csv(output_file, index=False)
        print(f"Daily PnL CSV saved to {output_file}")

        # Calculate additional stats using quant.analytics.performance
        returns_series = daily_agg['Daily_PnL_Pct']
        returns = returns_series.values
        n = len(returns)
        if n > 0:
            # Core metrics via quant module (rf=0 to match original behavior)
            annual_mean = returns.mean() * TRADING_DAYS_PER_YEAR
            annual_std = annualized_volatility(returns_series)
            annual_geo = annualized_return(returns_series)
            max_dd = max_drawdown(returns_series)
            sharpe = sharpe_ratio(returns_series, risk_free_rate=0.0)
            daily_agg['Year'] = daily_agg['Date'].dt.year
            yearly_returns = daily_agg.groupby('Year')['Daily_PnL_Pct'].apply(lambda x: np.prod(1 + x) - 1 if np.all(1 + x > 0) else np.nan)

            # Distribution stats via quant module
            kurt = return_kurtosis(returns_series)
            skw = return_skewness(returns_series)
            win_rate = compute_win_rate(returns_series) * 100  # Daily win rate as percentage
            max_win = returns.max()
            # Mean holding days from sim_df
            sim_df['Holding_Days'] = (sim_df['ExitDate'] - sim_df['EntryDate']).dt.days
            mean_holding_days = sim_df['Holding_Days'].mean()
            # Trade win rate: percentage of trades with positive PnL
            trade_win_rate = (sim_df['PnL_USD'] > 0).mean() * 100

            # Realized PnL per trade stats
            pnl_usd_values = sim_df['PnL_USD'].values
            if len(pnl_usd_values) > 0:
                pnl_series = pd.Series(pnl_usd_values)
                pnl_kurt = return_kurtosis(pnl_series)
                pnl_skw = return_skewness(pnl_series)
            else:
                pnl_kurt = np.nan
                pnl_skw = np.nan

            # SPY stats - download using YFinanceProvider
            min_date = daily_agg['Date'].min()
            max_date = daily_agg['Date'].max()
            try:
                provider = YFinanceProvider()
                spy_prices = provider.fetch_prices(
                    ['^GSPC'],
                    start=str(min_date.date()),
                    end=str((max_date + pd.Timedelta(days=1)).date()),
                )
                spy_returns_series = spy_prices['^GSPC'].pct_change().dropna()
                spy_returns = spy_returns_series.values
                if len(spy_returns) > 0:
                    spy_annual_mean = spy_returns.mean() * TRADING_DAYS_PER_YEAR
                    spy_annual_std = annualized_volatility(spy_returns_series)
                    spy_annual_geo = annualized_return(spy_returns_series)
                    spy_max_dd = max_drawdown(spy_returns_series)
                    spy_sharpe = sharpe_ratio(spy_returns_series, risk_free_rate=0.0)
                    spy_kurt = return_kurtosis(spy_returns_series)
                    spy_skw = return_skewness(spy_returns_series)
                    spy_win_rate = compute_win_rate(spy_returns_series) * 100
                    spy_max_win = spy_returns.max()
                    # SPY holding days not applicable, set to NaN
                    spy_mean_holding_days = np.nan
                    spy_trade_win_rate = np.nan  # No trades for SPY
                    # SPY yearly returns
                    spy_df = spy_prices.reset_index()
                    spy_df.columns = ['Date', 'Close']
                    spy_df['Date'] = pd.to_datetime(spy_df['Date'])
                    spy_df = spy_df.sort_values('Date')
                    spy_df['Daily_Return'] = spy_df['Close'].pct_change()
                    spy_df['Year'] = spy_df['Date'].dt.year
                    spy_yearly_returns = spy_df.groupby('Year')['Daily_Return'].apply(lambda x: np.prod(1 + x) - 1 if np.all(1 + x > 0) else np.nan)
                else:
                    spy_annual_mean = spy_annual_std = spy_annual_geo = spy_max_dd = spy_sharpe = spy_kurt = spy_skw = spy_win_rate = spy_max_win = spy_mean_holding_days = spy_trade_win_rate = np.nan
                    spy_yearly_returns = {}
            except Exception as e:
                print(f"Error downloading SPY data: {e}")
                spy_annual_mean = spy_annual_std = spy_annual_geo = spy_max_dd = spy_sharpe = spy_kurt = spy_skw = spy_win_rate = spy_max_win = spy_mean_holding_days = spy_trade_win_rate = np.nan
                spy_yearly_returns = {}

        # Create Excel with sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Daily PnL"
        # Write daily_agg to ws1
        for r, (date, pnl) in enumerate(zip(daily_agg['Date'], daily_agg['Daily_PnL_Pct']), 2):
            ws1.cell(row=r, column=1, value=date)
            ws1.cell(row=r, column=2, value=pnl)
        ws1.cell(1,1, 'Date')
        ws1.cell(1,2, 'Daily_PnL_Pct')

        # Sheet 2: Strategy Stats
        ws2 = wb.create_sheet("Strategy Stats")
        ws2.cell(1,1, 'Stat')
        ws2.cell(1,2, 'Value')
        ws2.cell(2,1, 'Annual Mean Return')
        ws2.cell(2,2, annual_mean)
        ws2.cell(3,1, 'Annual Stdev')
        ws2.cell(3,2, annual_std)
        ws2.cell(4,1, 'Annual Geometric Mean')
        ws2.cell(4,2, annual_geo)
        ws2.cell(5,1, 'Max Drawdown')
        ws2.cell(5,2, max_dd)
        ws2.cell(6,1, 'Sharpe Ratio')
        ws2.cell(6,2, sharpe)
        ws2.cell(7,1, 'Kurtosis')
        ws2.cell(7,2, kurt)
        ws2.cell(8,1, 'Skewness')
        ws2.cell(8,2, skw)
        ws2.cell(9,1, 'Daily Win Rate (%)')
        ws2.cell(9,2, win_rate)
        ws2.cell(10,1, 'Trade Win Rate (%)')
        ws2.cell(10,2, trade_win_rate)
        ws2.cell(11,1, 'Max Win')
        ws2.cell(11,2, max_win)
        ws2.cell(12,1, 'Mean Holding Days')
        ws2.cell(12,2, mean_holding_days)
        ws2.cell(13,1, 'PnL Kurtosis')
        ws2.cell(13,2, pnl_kurt)
        ws2.cell(14,1, 'PnL Skewness')
        ws2.cell(14,2, pnl_skw)
        # Return by year
        row = 15
        for year, ret in yearly_returns.items():
            ws2.cell(row,1, f'Return {year}')
            ws2.cell(row,2, ret)
            row += 1

        # Sheet 3: SPY Stats
        ws3 = wb.create_sheet("SPY Stats")
        ws3.cell(1,1, 'Stat')
        ws3.cell(1,2, 'Value')
        ws3.cell(2,1, 'Annual Mean Return')
        ws3.cell(2,2, spy_annual_mean)
        ws3.cell(3,1, 'Annual Stdev')
        ws3.cell(3,2, spy_annual_std)
        ws3.cell(4,1, 'Annual Geometric Mean')
        ws3.cell(4,2, spy_annual_geo)
        ws3.cell(5,1, 'Max Drawdown')
        ws3.cell(5,2, spy_max_dd)
        ws3.cell(6,1, 'Sharpe Ratio')
        ws3.cell(6,2, spy_sharpe)
        ws3.cell(7,1, 'Kurtosis')
        ws3.cell(7,2, spy_kurt)
        ws3.cell(8,1, 'Skewness')
        ws3.cell(8,2, spy_skw)
        ws3.cell(9,1, 'Daily Win Rate (%)')
        ws3.cell(9,2, spy_win_rate)
        ws3.cell(10,1, 'Trade Win Rate (%)')
        ws3.cell(10,2, spy_trade_win_rate)
        ws3.cell(11,1, 'Max Win')
        ws3.cell(11,2, spy_max_win)
        ws3.cell(12,1, 'Mean Holding Days')
        ws3.cell(12,2, spy_mean_holding_days)
        # Return by year for SPY
        row = 13
        for year, ret in spy_yearly_returns.items():
            ws3.cell(row,1, f'Return {year}')
            ws3.cell(row,2, ret)
            row += 1

        wb.save('pnl_stats.xlsx')
        print("Excel file with sheets saved as pnl_stats.xlsx")
    else:
        print("No data to process.")

# Call the function
calculate_daily_pnl()

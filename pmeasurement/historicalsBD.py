'''
Crear listado de activos. (Ponerlos todos)

1) Hacer excel que genera hojas por activo y en la primer
celda del activo pone la función de byma data de precios
históricos

2)Eliminar el apostrofe con el buscador de excel configurandolo
en libro

3)Eliminar la primera celda así no hay problema con la función

4)Crear un solo excel con los precios de los activos

5) Calcular dólar mep para luego pasar todo a USD

6) Buscar la manera automática de que la renta de los activos
se sume a los precios.

'''


import openpyxl

# Lista de activos

tickers = [
    "YPFD","GGAL","BBAR","SUPV","PAMP","TGSU2","TECO2","IRSA","CRESY","LOMA","BMA","EDN","CEPU",
    "AAL", "AAP", "AAPL", "ABBV", "ABEV", "ABNB", "ABT",
    "ACN", "ADBE", "AGRO", "ADI", "ADP", "AEM", "AMAT",
    "AMD", "AMGN", "AMX", "AMZN", "ANF", "ARCO", "ARM", "ASR",
    "AVGO", "AVY", "AZN","ASML","AI","TEAM","CLS","CEG","DECK","RGTI",
    "B", "BA", "BABA","NOW","VST","VRTX","PATH","PDD","XPEV",
    "BAK", "BB", "BHP","BRK-B",
    "BIDU", "BIIB", "BIOX", "BITF", "BKNG", "BKR", "BMY",
    "BP", "BRFS", "CAAP", "CAH",
    "CAR", "CAT", "CCL", "CDE", "CL", "COIN", "COST", "CRM",
    "CSCO", "CSNA3.SA", "CVS", "CVX", "CX", "DAL", "DD",
    "DE", "DEO", "DHR", "DIS", "DOCU", "DOW",
    "E", "EA", "EBAY", "EBR","EFX", "EQNR", "ERIC",
    "ERJ", "ETSY", "F", "FCX", "FDX",
    "FMX", "FSLR", "GE", "GFI", "GGB", "GILD",
    "GLOB", "GLW", "GM", "GOOGL", "GRMN",
    "GSK", "GT", "HAL", "HAPV3.SA", "HD", "HL", "HMC", "HMY",
    "HOG", "HON", "HPQ", "HSY", "HUT", "HWM",
     "IBM", "INFY", "INTC", "IP",
    "ISRG","JBSS3.SA",
    "JD", "JMIA", "JNJ","JOYY", "KEP", "KGC",
    "KMB", "KO", "KOD", "LAC", "LAR", "LLY",
    "LMT", "LND", "LRCX", "LREN3.SA", "LVS",  "MA", "MCD",
     "MDLZ", "MDT", "MELI", "META", "MGLU3.SA", "MMC", "MMM",
    "MO", "MOS", "MRK", "MRNA", "MRVL", "MSFT", "MSI", "MSTR",
    "MU", "MUX", "NEM", "NFLX", "NG", "NGG", "NIO",
    "NKE", "NTCO", "NTES", "NUE", "NVDA", "NVS",
    "NXE", "ORCL", "ORLY", "OXY", "PAAS", "PAC", "PAGS", "PANW", "PBI",
    "PBR", "PCAR", "PEP", "PFE", "PG",
    "PHG", "PINS", "PLTR", "PM", "PRIO3.SA", "PSX", "PYPL", "QCOM",
     "RACE", "RBLX", "RENT3.SA", "RIO", "RIOT", "ROKU", "ROST", "RTX",
     "SAP", "SATL", "SBS", "SBUX", "SCCO", "SDA", "SE",
    "SHEL", "SHOP", "SID", "SLB", "SNA", "SNAP", "SNOW", "SONY", "SPCE",
    "SPGI", "SPOT", "STLA", "STNE", "SYY", "T",
    "TCOM", "TEN", "TGT", "TIMB", "TM",
    "TMO", "TMUS", "TRIP", "TSLA", "TSM", "TTE", "TV", "TWLO",
    "TXN", "UAL", "UBER", "UGP", "UL", "UNH",
     "UNP", "URBN", "V", "VALE",
      "VIST", "VIV", "VOD", "VRSN", "VZ", "WBA", "WB", "WEGE3.SA", "WMT", "X", "XOM", "XP", "XRX",
    "XYZ", "YELP", "ZM", "ADS","AEG","AXP","AIG","BBD","BBAS3.SA","BSBR","SAN","BAC","BCS","BAS","BAYN",
    "BBV","SCHW","C","ELP","CS","BSN","DTEA","EOAN","AKO-B","FNMA", "FMCC","GPRK","HDB","HHPD",
    "HSBC","IBN","ING","IFF","JPM","ITUB","JCI","KB","LYG","MBG","MUFG","NEC1.HM","NSAN",
    "NOKA","NMR","PSO","PKS","SMSN","SWKS","TTM","RCTB4.BA","TIIAY","TEFO","TX","BK","GS","TRVV",
    "TJX", "USB", "UPST","WBO","WFC","AUY","YZCAY"
]

# Lista adicional a agregar (sin duplicados)
extra_tickers = '''AL30 AL30D AL29D AL35D AE38D AL41D GD29D GD30D GD35D GD38D GD41D BPA7D BPB7D BPC7D BPD7D BPJ5D BPY6D TX26 TZX26 TZXM6 TZXO6 TX28 TZXD6 TZXM7 TZX27 TZXD7 TZX28 DICP TX31 PARP CUAP S16Y5 S30Y5 S18J5 S30J5 S31L5 S15G5 S29G5 S12S5 S30S5 T17O5 S31O5 S10N5 S28N5 T15D5 T30E6 T13F6 T30J6 T15E7 TTM26 TTJ26 TTS26 TTD26 TLC5D NPCBD YMCHD MTCGD MGC9D PNDCD GNCXD RCCJD MRCAD IRCFD CAC5D BYCHD YMCID DNC7D RUCDD TLCMD TSC3D ARC1D MGCMD YMCXD PNXCD YFCJD TTCAD YMCJD YM34D MGCOD VSCTD SPY IWM QQQ IEUR VEA FXI EWZ EEM GLD XLRE YPFD GGAL BMA VIST MELI PAMP TGSU2 CEPU BYMA IBIT ETHA SLV VIG IJH EWJ PFE MO MRK SLB SCCO AZN BHP BIIB CAAP BKR KGC UAL YELP HPQ PBI TGT VZ AVGO FCX PYPL AAL ANF NEM DAL FSLR HAL GFI B OXY EQNR RIO UNH TLCPD TY30P S29Y6 PM GILD T RBLX PLTR KO UL BRKB JNJ PG NKE EA GLOB UL'''.split()

# Agregar solo los que no están en tickers
for t in extra_tickers:
    if t not in tickers:
        tickers.append(t)

# Ordenar alfabéticamente

# Eliminar posibles espacios y ordenar
unique_tickers = sorted(set(t.strip() for t in tickers if t.strip()))

# Fechas y parámetros fijos
fecha_inicio = "2025-01-01"
fecha_fin = "2025-07-10"
frecuencia = "Diario"
columna = "CORRIENTE"
periodo = "24"
precio = "PPT"

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()

for i, activo in enumerate(unique_tickers):
    # Crear o seleccionar la hoja
    if i == 0:
        ws = wb.active
        ws.title = activo
    else:
        ws = wb.create_sheet(title=activo)
    # Escribir la fórmula como texto en la celda A1 (con apóstrofe al inicio)
    formula = f"'=@SerieHistoricaPorEspecieBD(\"{activo}\";\"{frecuencia}\";\"{fecha_inicio}\";\"{fecha_fin}\";\"{columna}\";\"{periodo}\";\"{precio}\")"
    ws["A1"] = formula

# Guardar el archivo
wb.save("byma_activos.xlsx")

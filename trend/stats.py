import pandas as pd
import os
import numpy as np
import openpyxl

def calculate_daily_pnl(simulation_file='simulation_results.csv', main_file='main.csv', output_file='daily_pnl.csv'):
    # Load data
    sim_df = pd.read_csv(simulation_file, sep=';')  # Assuming semicolon separator
    main_df = pd.read_csv(main_file)
    
    # Convert dates with specified format
    sim_df['EntryDate'] = pd.to_datetime(sim_df['EntryDate'], format='%d/%m/%Y')
    sim_df['ExitDate'] = pd.to_datetime(sim_df['ExitDate'], format='%d/%m/%Y')
    main_df['Date'] = pd.to_datetime(main_df['Date'])  # Assuming main.csv dates are already parseable
    
    # Sort main_df for efficiency
    main_df = main_df.sort_values(['ticker', 'Date'])
    
    # Pre-group main_df by ticker for faster access
    ticker_data = {ticker: group for ticker, group in main_df.groupby('ticker')}
    
    daily_pnl_list = []
    total_trades = len(sim_df)
    
    for idx, trade in enumerate(sim_df.iterrows()):
        ticker = trade[1]['Ticker']  # trade is (index, row)
        entry_date = trade[1]['EntryDate']
        exit_date = trade[1]['ExitDate']
        usd_position = trade[1]['USD_Position']
        
        if ticker in ticker_data:
            # Filter the pre-grouped data for this ticker and date range, and valid reasons
            trade_days = ticker_data[ticker][
                (ticker_data[ticker]['Date'] >= entry_date) &
                (ticker_data[ticker]['Date'] <= exit_date) &
                (ticker_data[ticker]['Reason'].isin(['entry', 'hold', 'stop_loss', 'exit_min100']))
            ].sort_values('Date')
            
            if not trade_days.empty:
                # Calculate daily PnL contribution and weight
                closes = trade_days['Close'].values
                dates = trade_days['Date'].values
                for i in range(1, len(closes)):
                    daily_return = (closes[i] - closes[i-1]) / closes[i-1]
                    daily_pnl_contrib = usd_position * daily_return
                    daily_pnl_list.append({'Date': dates[i], 'Daily_PnL_Contrib': daily_pnl_contrib, 'Weight': usd_position})
        
        # Print progress
        progress = int((idx + 1) / total_trades * 100)
        print(f"Progress: {progress}%")
    
    # Aggregate by date: weighted average return
    if daily_pnl_list:
        daily_pnl_df = pd.DataFrame(daily_pnl_list)
        daily_agg = daily_pnl_df.groupby('Date').agg(
            Total_Contrib=('Daily_PnL_Contrib', 'sum'),
            Total_Weight=('Weight', 'sum')
        ).reset_index()
        daily_agg['Daily_PnL_Pct'] = daily_agg['Total_Contrib'] / daily_agg['Total_Weight']
        daily_agg = daily_agg[['Date', 'Daily_PnL_Pct']].sort_values('Date')
        daily_agg.to_csv(output_file, index=False)
        print(f"Daily PnL CSV saved to {output_file}")
        
        # Create Excel with sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Daily PnL"
        # Write daily_agg to ws1
        for r, (date, pnl) in enumerate(zip(daily_agg['Date'], daily_agg['Daily_PnL_Pct']), 2):
            ws1.cell(row=r, column=1, value=date)
            ws1.cell(row=r, column=2, value=pnl)
        ws1.cell(1,1, 'Date')
        ws1.cell(1,2, 'Daily_PnL_Pct')
        
        # Calculate stats
        returns = daily_agg['Daily_PnL_Pct'].values
        n = len(returns)
        if n > 0:
            # Annual mean return
            annual_mean = returns.mean() * 252
            # Annual stdev
            annual_std = returns.std() * np.sqrt(252)
            # Geometric mean annualized
            if np.all(1 + returns > 0):
                geo_mean_daily = np.prod(1 + returns) ** (1 / n) - 1
                annual_geo = (1 + geo_mean_daily) ** 252 - 1
            else:
                annual_geo = np.nan  # Handle negative products
            # Max drawdown
            cum_returns = np.cumprod(1 + returns)
            peak = np.maximum.accumulate(cum_returns)
            drawdown = (cum_returns - peak) / peak
            max_dd = drawdown.min()
            # Sharpe ratio (assuming 0 risk-free rate)
            sharpe = annual_mean / annual_std if annual_std > 0 else 0
            # Return by year
            daily_agg['Year'] = daily_agg['Date'].dt.year
            yearly_returns = daily_agg.groupby('Year')['Daily_PnL_Pct'].apply(lambda x: np.prod(1 + x) - 1 if np.all(1 + x > 0) else np.nan)
            
            # Sheet 2: Stats
            ws2 = wb.create_sheet("Stats")
            ws2.cell(1,1, 'Stat')
            ws2.cell(1,2, 'Value')
            ws2.cell(2,1, 'Annual Mean Return')
            ws2.cell(2,2, annual_mean)
            ws2.cell(3,1, 'Annual Stdev')
            ws2.cell(3,2, annual_std)
            ws2.cell(4,1, 'Annual Geometric Mean')
            ws2.cell(4,2, annual_geo)
            ws2.cell(5,1, 'Max Drawdown')
            ws2.cell(5,2, max_dd)
            ws2.cell(6,1, 'Sharpe Ratio')
            ws2.cell(6,2, sharpe)
            # Return by year
            row = 7
            for year, ret in yearly_returns.items():
                ws2.cell(row,1, f'Return {year}')
                ws2.cell(row,2, ret)
                row += 1
            
            wb.save('pnl_stats.xlsx')
            print("Excel file with sheets saved as pnl_stats.xlsx")
        else:
            print("No returns to calculate stats.")
    else:
        print("No data to process.")

# Call the function
calculate_daily_pnl()